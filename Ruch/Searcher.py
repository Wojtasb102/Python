from openpyxl import *
import os
import sqlite3


class Searcher:
    def __init__(self, excel_name, db_name):
        self.excel_name = excel_name
        self.db_name = db_name
        self.dir_name = os.path.dirname(__file__)
        self.db_dir = os.path.join(self.dir_name, self.db_name)
        self.workbook = load_workbook(filename=self.excel_name)
        self.worksheet = self.workbook.active
        self.dzialy = {"N-2": 2, "G-2": 3, "NE-1": 4, "NE-2": 5, "N-3": 6, "H-2": 7, "NK": 8, "Umysłowi": 9,
                       "NE (NE-2)": 5, "NE (NE-1)": 4}

    def create_db_connection(self):
        file_name = os.path.join(self.dir_name, self.db_name)
        self.conn = sqlite3.connect(file_name)
        self.c = self.conn.cursor()
        self.create_db_connection()

    def execute_query(self, query):
        conn = sqlite3.connect(self.db_dir)
        c = conn.cursor()
        c.execute(query)
        conn.commit()
        results = c.fetchall()
        conn.close()
        return results

    def get_cell(self, x, y):
        #     return self.worksheet.cell(row=x, column=y).number_format
        return self.worksheet.cell(row=y, column=x).value

    def set_cell(self, x, y, value):
        #     return self.worksheet.cell(row=x, column=y).number_format
        self.worksheet.cell(row=y, column=x).value = value
        #self.workbook.save(self.excel_name)

    def get_dimension(self):
        dimension = (self.worksheet.max_column, self.worksheet.max_row)
        return dimension

    def change_worksheet(self, sheet_name):
        self.worksheet = self.workbook[sheet_name]

    def get_date_index(self, date):
        # print ("SELECT Lp FROM Daty WHERE Data = {}".format(date))
        return self.execute_query(
            "SELECT Lp FROM Daty WHERE Data = '{}'".format(date))[0][0]

    def get_db_row(self, row_number, sheetname, column_name):
        return self.execute_query(
            "SELECT {} FROM {} WHERE Lp = '{}'".format(column_name, sheetname, row_number))[0]

    def calculate_changes(self, sheetname, operation):
        for x in range(1, self.execute_query("SELECT COUNT (Lp) FROM {} ".format(sheetname))[0][0] + 1):
            query_response = self.get_db_row(x, sheetname, "Dział, Data")
            print(query_response[0])
            print(query_response[1])
            self.change_worksheet("inne")
            column = self.dzialy[query_response[0]]
            row = int(self.get_date_index(query_response[1]))
            print("kolumna: {}".format(column))

            self.write(column, row, operation)

    def calculate_transfer(self, sheetname):
        print(self.execute_query("SELECT COUNT (Lp) FROM {} ".format(sheetname))[0][0] + 1)
        for x in range(1, self.execute_query("SELECT COUNT (Lp) FROM {} ".format(sheetname))[0][0] + 1):
            query_response = self.get_db_row(x, sheetname, "Dział, Data,[Dział poprzedni]")
            print(query_response[0])
            print(query_response[1])
            print(query_response[2])
            self.change_worksheet("inne")
            row = self.get_date_index(query_response[1])
            column_to = self.dzialy[query_response[0]]
            column_from = self.dzialy[query_response[2]]
            print("kolumna: {}".format(column_to))

            self.write(column_to, row, 1)
            self.write(column_from, row, -1)

    def write(self, column, row, operation):
        present_value = self.get_cell(column, row)
        if present_value is None:
            present_value = 0
        self.set_cell(column, row, present_value + (1 * operation))

    def save(self):
        self.workbook.save(self.excel_name)

    def clear(self):
        self.change_worksheet("inne")
        print(self.get_dimension())
        max_x = self.get_dimension()[0]+1
        max_y = self.get_dimension()[1]+1
        for x in range (2, max_x):
            for y in range (2, max_y):
                #self.set_cell(x, y, None)
                self.worksheet.cell(row=y, column=x).value = None
