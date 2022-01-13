from openpyxl import *
import os
import sqlite3


class Connector:
    def __init__(self, excelName, dbName):
        self.excelName = excelName
        self.dbName = dbName
        self.dirname = os.path.dirname(__file__)
        if not self.is_file(self.excelName):
            workbook = Workbook()
            workbook.save(filename=excelName)
        self.workbook = load_workbook(filename=self.excelName)
        self.worksheet = self.workbook.active

        self.db_dir = os.path.join(self.dirname, self.dbName)
        self.create_db_connection()

        # if not self.is_file(self.dbName):
        #     filename = os.path.join(self.dirname, self.dbName)
        #     conn = sqlite3.connect(filename)

    def is_file(self, fileName):
        # dirname = os.path.dirname(__file__)
        filename = os.path.join(self.dirname, fileName)
        print(os.path.isfile(filename))
        return os.path.isfile(filename)

    def sheet_list(self):
        return self.workbook.sheetnames

    def get_cell(self, x, y):
        #     return self.worksheet.cell(row=x, column=y).number_format
        return self.worksheet.cell(row=y, column=x).value

    def get_type(self, x, y):
        return self.worksheet.cell(row=y, column=x).number_format

    def change_worksheet(self, sheet_name):
        self.worksheet = self.workbook[sheet_name]

    def get_dimension(self):
        dimension = (self.worksheet.max_column, self.worksheet.max_row)
        return dimension

    def get_max_x(self):
        return self.worksheet.max_column

    def get_max_y(self):
        return self.worksheet.max_row

    def get_header(self):
        header = []
        for i in range(1, self.get_max_x() + 1):
            header.append(self.get_cell(i, 1))
        return header

    def get_header_sql_type(self):
        header = self.get_header_type()
        typy = {"@": "Varchar(30)", "0": "Integer", "0.00": "Integer", "General": "Ogólne"}
        typsy = []
        for x in header:
            if not (x in typy.keys()):
                typsy.append("Varchar(30)")
            else:
                typsy.append(typy[x])
        return typsy

    def get_header_type(self):
        header_type = []
        for i in range(1, self.get_max_x() + 1):
            header_type.append(self.get_type(i, 2))
        return header_type

    # Do ewentualnej poprawy: Dodawanie nawiasów w przypadku typu TEKSTOWEGO
    def get_row(self, row_number):
        row = []
        for column in range(1, self.get_max_x() + 1):
            if self.get_header_type()[column - 1] == "@":
                row.append("'" + str(self.get_cell(column, row_number)) + "'")
            else:
                row.append(str(self.get_cell(column, row_number)))
        # print(row)
        return row

    # ----------------------------- DB METODH

    def create_db_connection(self):
        filename = os.path.join(self.dirname, self.dbName)
        self.conn = sqlite3.connect(filename)
        self.c = self.conn.cursor()

    #    def create_db_connection2(self):

    def create_example_tables(self):
        if 'NazwaTabeli' not in self.read_table_names():
            self.execute_query(
                "CREATE TABLE`NazwaTabeli`('Pole1' INTEGER,'Pole2'VARCHAR(20),'Pole3'INTEGER,'Field4'INTEGER);")
        else:
            print("Tabele 1 nie tworzona")
        if 'NazwaTabeli2' not in self.read_table_names():
            self.execute_query(
                "CREATE TABLE`NazwaTabeli2`('Pole21'INTEGER,'Pole22'VARCHAR(20),'Pole23'INTEGER,'Field24'INTEGER);")
        else:
            print("Tabele 2 nie tworzona")

    # self.conn.close()

    def delete_all_tables(self):
        for tablenames in self.read_table_names():
            print("DROP table '{}';".format(tablenames))
            querry = "DROP table '{}';".format(tablenames)
            # self.c.execute(querry)
            self.execute_query(querry)

    def read_table_names(self):
        # self.c.execute("SELECT name FROM sqlite_master WHERE type='table'")
        table_names = []
        results = self.execute_query("SELECT name FROM sqlite_master WHERE type='table'")
        for position in results:
            table_names.append(position[0])

        return table_names

    def execute_query(self, query):
        conn = sqlite3.connect(self.db_dir)
        c = conn.cursor()
        c.execute(query)
        conn.commit()
        results = c.fetchall()
        conn.close()
        return results

    def print_data(self):
        i = 1
        for x in self.get_header():
            for y in range(2, self.get_max_y() + 1):
                print(self.get_cell(i, y))
            i = i + 1

    def excel_into_sql(self):
        self.create_table(self.sheet_list()[0])

    def create_table(self, sheet_name):
        fields = []
        for i, x in enumerate(self.get_header()):
            fields.append("'{}'{}".format(x, self.get_header_sql_type()[i]))
        a = ','.join(fields)
        print(a)
        query = "CREATE TABLE '{}' (".format(sheet_name)
        query += (a + ");")
        print(query)
        self.execute_query(query)
        # "CREATE TABLE`NazwaTabeli`('Pole1' INTEGER,'Pole2'VARCHAR(20),'Pole3'INTEGER,'Field4'INTEGER);")

    def insert_all_excel_data_into_sql(self):
        for sheet in self.sheet_list():
            self.change_worksheet(sheet)
            self.create_table(sheet)
            self.insert_sheet_into_sql(sheet)


    # INSERT
    # INTO
    # table(column1, column2,..)
    # VALUES(value1, value2, ...);
    def insert_row_from_excel_to_sql(self, row, sheet_name):
        head = []
        value = []
        x = self.get_row(row)
        value = ",".join(x)
        query = "INSERT INTO '{}'VALUES({});".format(sheet_name, value)
        print(query)
        # print(value)
        self.execute_query(query)

    def insert_sheet_into_sql(self, sheet):
        for i in range (2, self.get_max_y() + 1):
            self.insert_row_from_excel_to_sql(i, sheet)


def main():
    print("Start Main")
    conector = Connector("Excel.xlsx", "Database.db")
    conector.is_file("Excel.xlsx")
    conector.is_file("Database.db")
    print("Start Main")
    print(conector.sheet_list())
    conector.change_worksheet(conector.sheet_list()[0])
    print(conector.get_cell(1, 1))
    conector.change_worksheet(conector.sheet_list()[1])
    print(conector.get_cell(1, 1))
    conector.change_worksheet(conector.sheet_list()[0])
    print(conector.get_dimension()[0])
    print(conector.get_header())
    conector.print_data()
    print("Wymiar {}".format(conector.get_dimension()))
    print("Wymiar {}, {}".format(conector.get_max_x(), conector.get_max_y()))
    print(conector.get_header())
    print(conector.get_header_type())
    typy = {"@": "Varchar", "0": "Integer", "0.00": "Integer", "General": "Ogólne"}
    typsy = []
    for x in conector.get_header_type():
        if not (x in typy.keys()):
            typsy.append("Varchar")
        else:
            typsy.append(typy[x])
    print(",".join(typsy))
    print(conector.read_table_names())
    # conector.create_example_tables()
    conector.delete_all_tables()
    print("DANE:")
    # for i in range(1, conector.get_max_y() + 1):
    #     conector.get_row(i)
    # print(conector.get_row(2))
    #conector.create_table("Sheet")
    print(conector.insert_all_excel_data_into_sql())

    # print(conector.get_header_sql_type())


main()
